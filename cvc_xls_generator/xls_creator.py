## required imports
import re
import os
import sys
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

class GenerateXls:

    def __init__(self, json_file, output_file, comments_file=None, escaped_path=[]):
        self.remove_error = re.compile(''',"analysisExceptions":\[\{<exception>.*</exception>\}\]''')
        self.dependency_check_json = json_file
        self.output_file_name = output_file
        self.escaped_path = escaped_path
        self.comments_file = comments_file

    def readCVCDataFromJsonFile(self):
        if os.path.isfile(self.dependency_check_json):
            try:
                with open(self.dependency_check_json, 'r') as f:
                    data = f.read()
                    if re.search(self.remove_error, data):
                        data = re.sub(self.remove_error, '', data)
                cvcJsonData = json.loads(data)
            except Exception as e:
                sys.exit(e)
        else:
            sys.exit("dependency check json report not present at" + str(self.dependency_check_json))
        return cvcJsonData

    def readCommentsFile(self):
        if self.comments_file:
            if os.path.isfile(self.comments_file):
                try:
                    with open(self.comments_file, 'r') as rb:
                        comments_data = json.loads(rb.read())
                except Exception as e:
                    sys.exit('invalid comments file '+str(e))
            else:
                sys.exit("Comments file Not present")
        else:
            comments_data = {}
        return comments_data

    def cvcJsonDataToDataFrame(self):
        vuldependency, cve_id, severity, filePath, description, status, auditor_comments = [[] for i in range(7)]
        try:
            cvcJsonData = self.readCVCDataFromJsonFile()
            comments_data = self.readCommentsFile()
            for dependency in cvcJsonData.get("dependencies", {}):
                if 'vulnerabilities' in dependency.keys():
                    for vulnerability in dependency.get('vulnerabilities', {}):
                        if 'relatedDependencies' in dependency.keys():

                            vuldependency.append(dependency.get('fileName').strip())
                            cve_id.append(vulnerability.get('name').strip())
                            severity.append(vulnerability.get('severity').upper().strip())
                            filePath.append('/'.join([x for x in dependency.get('filePath').split("\\") if x not in self.escaped_path]))
                            description.append(str(vulnerability.get('description')).replace('\n', ' '))
                            status.append(comments_data.get(vuldependency[-1], {}).get(cve_id[-1], {}).get("Status", "Open"))
                            auditor_comments.append(comments_data.get(vuldependency[-1], {}).get(cve_id[-1], {}).get("Comment", "issue details not present in comments json file"))

                            for relatedDependency in dependency.get('relatedDependencies', {}):
                                filename = relatedDependency.get('filePath').split('\\')[-1].strip()
                                filePath.append('/'.join([x for x in relatedDependency.get('filePath').split('\\') if x not in self.escaped_path]))
                                vuldependency.append(filename)
                                cve_id.append(vulnerability.get('name').strip())
                                description.append(str(vulnerability.get('description')).replace('\n', ' '))
                                severity.append(vulnerability.get('severity').upper().strip())
                                status.append(comments_data.get(vuldependency[-1], {}).get(cve_id[-1], {}).get("Status", "Open"))
                                auditor_comments.append(comments_data.get(vuldependency[-1], {}).get(cve_id[-1], {}).get("Comment", "issue details not present in comments json file"))
                        else:
                            vuldependency.append(dependency.get('fileName').strip())
                            cve_id.append(vulnerability.get('name').strip())
                            severity.append(vulnerability.get('severity').upper())
                            filePath.append('/'.join([x for x in dependency.get('filePath').split('\\') if x not in self.escaped_path]))
                            description.append(str(vulnerability.get('description')).replace('\n', ' '))
                            status.append(comments_data.get(vuldependency[-1], {}).get(cve_id[-1], {}).get("Status", "Open"))
                            auditor_comments.append(comments_data.get(vuldependency[-1], {}).get(cve_id[-1], {}).get("Comment", "issue details not present in comments json file"))

            result_data = zip(vuldependency, description, cve_id, severity,filePath, status, auditor_comments)
            df_cvc = pd.DataFrame(list(result_data),
                columns = [
                    "DependencyName",
                    "Description",
                    "CVE",
                    "Severity",
                    "FilePath",
                    "Status",
                    "Auditor Comment"
                ]
            )
        except Exception as e:
            df_cvc = None
            sys.exit(e)

        return df_cvc

    def makeXL(self):
        try:
            df_cvc = self.cvcJsonDataToDataFrame()
            workbook = Workbook()
            workbook.remove(workbook.active)
            header_font = Font(name='Calibri',bold=True,color='FFFFFF')
            centered_alignment = Alignment(horizontal='center')
            wrapped_alignment = Alignment(vertical='top',wrap_text=False)
            fill = PatternFill(start_color='5FABE6',end_color='5FABE6',fill_type='solid',)
            cvc_sheet_columns = [
                ('DependencyName', 40),
                ('Description', 40),
                ('CVE', 30),
                ('Severity', 15),
                ('FilePath', 40),
                ('Status', 15),
                ('Auditor Comment', 40),
                ('Developer Comment', 40)
            ]
            worksheet = workbook.create_sheet(title='CVC',index=0)
            row_num = 1
            for col_num, (column_title, column_width) in enumerate(cvc_sheet_columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.alignment = centered_alignment
                cell.fill = fill
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = column_width
            for i in range(len(df_cvc['DependencyName'])):
                row_num += 1
                row = [
                    (df_cvc.loc[i,'DependencyName'],'Normal'),
                    (df_cvc.loc[i,'Description'],'Normal'),
                    (df_cvc.loc[i,'CVE'],'Normal'),
                    (df_cvc.loc[i,'Severity'],'Normal'),
                    (df_cvc.loc[i,'FilePath'],'Normal'),
                    (df_cvc.loc[i,'Status'],'Normal'),
                    (df_cvc.loc[i,'Auditor Comment'],'Normal')
                ]
                for col_num, (cell_value, cell_format) in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.style = cell_format
                    cell.alignment = wrapped_alignment
            worksheet.freeze_panes = worksheet['A2']
            worksheet.sheet_properties.tabColor = '5FABE6'
            workbook.save(self.output_file_name)
            print()
            print('execl created successfully....')
            print()
            return
        except Exception as e:
            print("Unable to create xls....")
            sys.exit(e)
